import sys
import pandas as pd
import numpy as np
import decimal

mileageRate = float(sys.argv[1])
summary = open("summary.xlsx", "rb")
a = pd.read_excel(summary, sheet_name="Entries")

#-Summarize per week-#
weeklySummary = a.groupby(["Company", "Name", "Rate", "Multiplier", "Week"], sort=False)[["Hours", "Miles"]].sum(numeric_only=True)
companyIndex = a.groupby(["Company"], sort=False).count().index.values
weeklySummary = weeklySummary.reindex(companyIndex, level="Company")
regHours = np.where(weeklySummary["Hours"] > 40, 40, weeklySummary["Hours"])
otHours = np.where(weeklySummary["Hours"] > 40, weeklySummary["Hours"] - 40, 0)
weeklySummary.insert(1, "Reg Hours", regHours)
weeklySummary.insert(2, "OT Hours", otHours)

#-Summarize per month-#
monthlySummary = weeklySummary.droplevel(["Week"]).groupby(["Company", "Name", "Rate", "Multiplier"], sort=False).sum(numeric_only=True)
monthlySummary = monthlySummary.reset_index(level=["Rate", "Multiplier"])
monthlySummary["Hours Total"] = round(monthlySummary["Hours"] * monthlySummary["Rate"], 2) * monthlySummary["Multiplier"]
#temp = (monthlySummary["Hours Total"] * 100).to_frame()
#print(type(temp))
#temp.iloc[:, 0] = temp.iloc[:, 0].apply(
#  lambda x: decimal.Decimal(x).to_integral_value(rounding=decimal.ROUND_HALF_UP)
#)
#monthlySummary["Hours Total"] = (temp / 100)
#monthlySummary["Hours Total"] = monthlySummary["Hours Total"] * monthlySummary["Multiplier"]

monthlySummary["Reg Hours Total"] = round(monthlySummary["Reg Hours"] * monthlySummary["Rate"], 2) * monthlySummary["Multiplier"]
monthlySummary["OT Hours Total"] = round(monthlySummary["OT Hours"] * monthlySummary["Rate"] * 1.5, 2) * monthlySummary["Multiplier"]
monthlySummary["OT Premium"] = round(monthlySummary["OT Hours"] * monthlySummary["Rate"] * 0.5, 2) * monthlySummary["Multiplier"]
monthlySummary["Miles Total"] = monthlySummary["Miles"] * mileageRate
monthlySummary["Total"] = monthlySummary["Reg Hours Total"] + monthlySummary["OT Hours Total"] + monthlySummary["Miles Total"]
monthlySummary = monthlySummary.drop(columns=["Multiplier"])

#-Summarize per company-#
companySummary = monthlySummary.drop(columns=["Rate"]).groupby(["Company"], sort=False).sum(numeric_only=True).reindex(companyIndex, level="Company")
fees = []
for i in companyIndex:
    try:
        fees.append(float(input(f"Enter fees for {i}: ")))
    except:
        fees.append(0)
companySummary.insert(8, "Fees", fees)
companySummary["Total"] += companySummary["Fees"]
companySummary.insert(8, "Total-Miles",companySummary["Total"] - companySummary["Miles Total"])

#-Totals-#
totals = companySummary.sum(numeric_only=True)

#-Distribute OT premium per jobs per week-#
a["Total"] = (a["Hours"] * a["Rate"] * a["Multiplier"]) + (mileageRate * a["Miles"])
weeklyCount = a.groupby(["Name", "Week"], sort=False)["Open"].count().reset_index().rename(columns={"Open":"Divisor"})
otEntries = weeklySummary[weeklySummary["OT Hours"] > 0].drop(columns=["Hours", "Reg Hours", "Miles"]).reset_index()
otEntries["Adder"] = (otEntries["Rate"] * 0.5 * otEntries["Multiplier"]) * otEntries["OT Hours"]
otEntries = otEntries.drop(columns=["Rate", "OT Hours"])
otEntries = otEntries.merge(weeklyCount, how='left')
otEntries["OT Adder"] = otEntries["Adder"] / otEntries["Divisor"]
otEntries = otEntries.drop(columns=["Adder", "Divisor"])
a = a.merge(otEntries, how='left').assign(Total=lambda d: d["Total"].add(d.pop("OT Adder"), fill_value=0))
print("Job Summary Total: ", a["Total"].sum() + sum(fees))

#-Calculate Job Summary-#
jobSummary = a.drop(columns=["Rate", "Multiplier"]).groupby(["Job", "Fund", "Activity", "Facility"]).sum(numeric_only=True)
jobSummary.loc[jobSummary["Open"] > 0, "Open"] = "Open"
jobSummary.loc[jobSummary["Open"] == 0, "Open"] = "Closed"
jobSummary.loc[("Fees", 6, 75041, 6000), :] = ("Closed", 0, 0, totals["Fees"])
jobSummary.set_index(["Open", "Hours", "Miles", "Total"])

#-Set Total per Job Summary-#
totals["Total"] = jobSummary["Total"].sum()

#-Stylize-#
style = {"border": "1px solid black", "background-color": "#abcdef"}
a = a.style.set_properties(**style)
weeklySummary = weeklySummary.style.set_properties(**style)
monthlySummary = monthlySummary.style.set_properties(**style)
companySummary = companySummary.style.set_properties(**style)
totals = totals.to_frame().style.set_properties(**style)
jobSummary = jobSummary.drop(columns=["Open", "Hours", "Miles"]).style.set_properties(**style)
def highlite(s):
    return ["background-color: #fff999" if s_ else None for s_ in s]
weeklySummary = weeklySummary.apply(highlite, subset=["OT Hours"], axis=1)
monthlySummary = monthlySummary.apply(highlite, subset=["OT Hours", "OT Hours Total", "OT Premium"], axis=1)

#-Write to summary.xlsx-#
with pd.ExcelWriter(
    "summary.xlsx",
    mode="a",
    engine="openpyxl",
    if_sheet_exists="replace",
) as writer:
    a.to_excel(writer, sheet_name="Entries", index=False)
    weeklySummary.to_excel(writer, sheet_name="Weekly")
    monthlySummary.to_excel(writer, sheet_name="Monthly")
    companySummary.to_excel(writer, sheet_name="CompanySummary")
    totals.to_excel(writer, sheet_name="Totals", header=False)
    jobSummary.to_excel(writer, sheet_name="JobSummary")

with pd.ExcelWriter("summary.xlsx", mode="a", if_sheet_exists="overlay") as writer:
    totals.to_excel(writer, sheet_name="JobSummary", startrow=writer.sheets["JobSummary"].max_row + 1, header=False)

try:
        from openpyxl.cell import get_column_letter
except ImportError:
        from openpyxl.utils import get_column_letter
        from openpyxl.utils import column_index_from_string
from openpyxl import load_workbook
import openpyxl
from openpyxl import Workbook


workbook = load_workbook("summary.xlsx")
for sheet_name in workbook.sheetnames:
  for column_cells in workbook[sheet_name].columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        if new_column_length == 17:
              new_column_length = 12
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            workbook[sheet_name].column_dimensions[new_column_letter].width = new_column_length*1.23
workbook.save("summary.xlsx")